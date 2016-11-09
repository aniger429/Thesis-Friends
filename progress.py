from openpyxl import load_workbook
from openpyxl import Workbook


# used openpyxl#
wb = load_workbook(filename = 'Test.xlsx') #input
sheet = wb['Sheet1']
rown = 2
column = 'A'
ti = 0
id = 1000001;
checkList = []
checkList2 = []
#output = open("cleaned_tweets.csv", "w")
userList = dict()
type(userList)
userFile = open("C:/Python27/anonymized.txt", 'w');
for row in sheet.rows:
    if (sheet.max_row < rown):
        break;
    else:
        d = sheet['A' + str(rown)]
        e = sheet['B' + str(rown)]
        f = sheet['D' + str(rown)]
        checkList.append(f.value)
        aa = e.value
        #aa = aa.encode(aa, 'ignore').decode('ascii')
        rown += 1
        if "@" in aa:
            splitList = aa.split()
            for i in range(0, len(splitList)):
                if "@" in splitList[i]:
                    if (len(splitList[i].replace("@", "")) > 0):
                        checkList.append(splitList[i].replace("@", ""))
checkList = list(set(checkList))
print(checkList)
for i in range(0, len(checkList)):
    checkList[i] = checkList[i].lower()
    userFile.write(checkList[i] + " " + str(i + id) + "\n")
userFile.close()

rown = 2
for row in sheet.rows:
    if (sheet.max_row < rown):
        break;
    else:
        e = sheet['B' + str(rown)]
        aa = e.value
        print("Original : \n" + aa + "\n")
        rown += 1
        if "@" in aa:
            splitList = aa.split()
            for i in range(0, len(splitList)):
                if "@" in splitList[i]:
                    splitList[i] = splitList[i].lower()
                    userFile = open("C:/Python27/anonymized.txt", 'r');
                    userFcount = len(userFile.readlines())
                    userFile.close();
                    userFile = open("C:/Python27/anonymized.txt", 'r');
                    for j in range(0, userFcount):
                        idRead = userFile.readline()
                        idSplice = idRead.split()
                        if splitList[i].replace("@", "") == idSplice[0]:
                            splitList[i] = "@" + idSplice[1]
            userFile.close()                
            aa = ""
            for i in range(0, len(splitList)):
                aa += " " + splitList[i]
                #User Anonymization by far#
        aa = aa.lower() #to lowercase#
        splitList = aa.split()
        for i in range(0, len(splitList)):
            standard = splitList[i]
            stopF = open("C:/Python27/stopwords.txt");
            contF = open("C:/Python27/contractions.txt");
            engF = open("C:/Python27/english.txt");
            Fcount = len(stopF.readlines())
            Fcount2 = len(contF.readlines())
            engFcount = len(engF.readlines())
            stopF.close();
            contF.close();
            engF.close();
            stopF = open("C:/Python27/stopwords.txt");
            contF = open("C:/Python27/contractions.txt");
           # engF = open("C:/Python27/english.txt");

                    

                
            for j in range(0, Fcount):       
                stopW = stopF.readline()
                stopW = stopW.strip()
                if stopW == splitList[i]:
                    splitList[i] = "" #stopword            
            for j in range(0, Fcount2):
                cont = contF.readline()
                contWords = cont.split()
                if contWords[0] == splitList[i]:
                    if (len(contWords) > 2):
                        splitList[i] = contWords[1] + " " + contWords[2]
                    else:
                        splitList[i] = contWords[1] #contraction#
                        
            if "http://" in splitList[i] or "https://" in splitList[i]: #link#
                splitList[i] = ""
            if "#" in splitList[i]: #Hashtag#
                splitList[i] = ""
            if "&" in splitList[i] and ";" in splitList[i]: #escape HTML char#
                splitList[i] = ""
            if "." in splitList[i]:
                splitList[i] = splitList[i].replace(".", "")
            if "," in splitList[i]:
                splitList[i] = splitList[i].replace(",", "")
            if "!" in splitList[i]:
                splitList[i] = splitList[i].replace("!", "")
            if "?" in splitList[i]:
                splitList[i] = splitList[i].replace("?", "")
            if ";" in splitList[i]:
                splitList[i] = splitList[i].replace(";", "")
            if ":" in splitList[i]:
                splitList[i] = splitList[i].replace(":", "")
            if "'" in splitList[i]:
                splitList[i] = splitList[i].replace("'", "")
            if "`" in splitList[i]:
                splitList[i] = splitList[i].replace("`", "")
            if "\"" in splitList[i]:
                splitList[i] = splitList[i].replace("\"", "")
            if "-" in splitList[i]:
                splitList[i] = splitList[i].replace("-", "")
            if "/" in splitList[i]:
                splitList[i] = splitList[i].replace("/", "")
            if "(" in splitList[i]:
                splitList[i] = splitList[i].replace("(", "")
            if "[" in splitList[i]:
                splitList[i] = splitList[i].replace("[", "")
            if ")" in splitList[i]:
                splitList[i] = splitList[i].replace(")", "")
            if "]" in splitList[i]:
                splitList[i] = splitList[i].replace("]", "")
            if "*" in splitList[i]:
                splitList[i] = splitList[i].replace("*", "")
            if "<" in splitList[i]:
                splitList[i] = splitList[i].replace("<", "")
            if ">" in splitList[i]:
                splitList[i] = splitList[i].replace(">", "")
            if "{" in splitList[i]:
                splitList[i] = splitList[i].replace("{", "")
            if "}" in splitList[i]:
                splitList[i] = splitList[i].replace("}", "")           
                



            for j in range(0, len(standard)-2):
                if (standard[j].replace("@", "").isdigit()):
                    continue
                elif (standard[j] == standard[j+1] == standard[j+2]):
                        splitList[i] = splitList[i].replace(standard[j+2], "", 1) #standardize#
            stopF.close()
            contF.close()

            #start of spliting attached words, this is not working properly
            #forsplit = 0
            #slicer = splitList[i]
            #slicedList = []
            #for j in range(0, engFcount):
      #          engl = engF.readline()
       #         engl = engl.strip()
        #        s = 0
         #       if len(engl) > 3:
         #           for k in range(1, len(splitList[i])):
          #              if engl == slicer[s:k]:
           #                 if (k + 1 <= len(splitList[i])):
            #                    s = k + 1
             #                   k = s + 1
              #                  forsplit += 1
               #                 slicer = slicer.replace(engl, "", 1)
                #                slicedList.append(engl)                
         #   if forsplit > 1:
          #      splitList[i] = ""
           #     for j in range(0, len(slicedList)):
            #        if (j == 0):
             #           splitList[i] = splitList[i] + slicedList[j]
              #      else:
               #         splitList[i] = splitList[i] + " " + slicedList[j]
         #   engF.close()
            #end of spliting            
                
                            
                          
        aa = ""

       
        for i in range(0, len(splitList)):
            aa += " " + splitList[i]
            #Removal of hashtags and link without exporting json file#
            #What would be the range of the emoticons? Do we need something like dictionary for emoticons as well? I need help with it#     
        print("After : \n " + aa + "\n")
        #output.write("%s\n" % aa)
#output.close()
        # print(slicedList)
            #split attach words to be done..#
            

        
                            

